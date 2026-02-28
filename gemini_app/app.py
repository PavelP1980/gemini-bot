from flask import Flask, render_template, request, jsonify
from dotenv import load_dotenv
import os
import google.generativeai as genai
import PyPDF2
import docx
import openpyxl

# ===== НАСТРОЙКИ ЭКОНОМИИ КВОТЫ =====
USE_GEMINI = True              # если поставить False, бекенд работает без запросов к модели
MAX_CONTEXT_CHARS = 8000       # максимум символов из документа, которые отправляем в модель
MAX_OUTPUT_TOKENS = 1024       # ограничение длины ответа модели

# Загрузка переменных окружения из .env
load_dotenv()

app = Flask(__name__)

# Загрузка ключа API из переменной окружения
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

# Проверка наличия API ключа
if not GEMINI_API_KEY:
    raise ValueError("GEMINI_API_KEY не найден в переменных окружения. Пожалуйста, установите его.")

# Настройка Gemini API
genai.configure(api_key=GEMINI_API_KEY)
model = genai.GenerativeModel(
    "gemini-2.5-flash",
    generation_config={"max_output_tokens": MAX_OUTPUT_TOKENS}
)

def extract_text_from_pdf(file_stream):
    reader = PyPDF2.PdfReader(file_stream)
    text = ""
    for page_num in range(len(reader.pages)):
        text += reader.pages[page_num].extract_text() or ""
    return text

def extract_text_from_docx(file_stream):
    document = docx.Document(file_stream)
    text = "\n".join([paragraph.text for paragraph in document.paragraphs])
    return text

def extract_text_from_xlsx(file_stream):
    workbook = openpyxl.load_workbook(file_stream)
    text = ""
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        text += f"--- Sheet: {sheet_name} ---\n"
        for row in sheet.iter_rows():
            row_values = [str(cell.value or "") for cell in row]
            text += "\t".join(row_values) + "\n"
    return text

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/ask_gemini", methods=["POST"])
def ask_gemini():
    user_input = request.form.get("text_input")
    file = request.files.get("file")

    if not user_input and not file:
        return jsonify({"response": "Пожалуйста, введите ваш вопрос или загрузите документ."}), 400

    context = ""
    if file:
        filename = file.filename.lower()
        if filename.endswith(".pdf"):
            context = extract_text_from_pdf(file.stream)
        elif filename.endswith((".doc", ".docx")):
            context = extract_text_from_docx(file.stream)
        elif filename.endswith((".xls", ".xlsx")):
            context = extract_text_from_xlsx(file.stream)
        elif filename.endswith((".png", ".jpg", ".jpeg")):
            # Для изображений, читаем бинарные данные напрямую
            file.stream.seek(0)
            image_data = file.stream.read()
            # Gemini API принимает изображения в виде словаря с mime_type и data
            context = {
                'mime_type': file.mimetype,
                'data': image_data
            }
        elif filename.endswith(".txt"):
            context = file.stream.read().decode('utf-8')
        else:
            return jsonify({"response": "Неподдерживаемый формат файла."}), 400

        # Обрезаем контекст, чтобы не слать гигантские документы
        if len(context) > MAX_CONTEXT_CHARS:
            context = context[:MAX_CONTEXT_CHARS]

    if context:
        full_prompt = (
            "Ты помощник по анализу документов. "
            "Отвечай кратко и по делу.\n\n"
            f"Документ:\n{context}\n\n"
            f"Вопрос пользователя: {user_input}"
        )
    else:
        full_prompt = user_input

    # Режим разработки без реальных запросов к модели
    if not USE_GEMINI:
        return jsonify({"response": "[DEBUG] Здесь был бы ответ от Gemini"}), 200

    try:
        if isinstance(context, dict) and 'mime_type' in context:
            # Мультимодальный запрос с изображением
            # Для изображений, 'context' уже содержит словарь с mime_type и data
            # Прим. для генерации content, требуется либо текст либо список из текста и картинки
            # Если user_input пуст, то отправляем только изображение с неявным запросом
            # Если user_input не пуст, то отправляем явный запрос и изображение
            if user_input:
                response = model.generate_content([user_input, context])
            else:
                response = model.generate_content([context])
        else:
            # Текстовый запрос (с документом или без)
            if context:
                full_prompt = (
                    "Ты помощник по анализу документов. "
                    "Отвечай кратко и по делу.\n\n"
                    f"Документ:\n{context}\n\n"
                    f"Вопрос пользователя: {user_input}"
                )
            else:
                full_prompt = user_input
            response = model.generate_content(full_prompt)

        response_text = response.text
    except Exception as e:
        print(f"Ошибка при обращении к Gemini API: {e}")
        response_text = f"Произошла ошибка при получении ответа от Gemini: {e}"

    return jsonify({"response": response_text})

if __name__ == "__main__":
    app.run(debug=True, port=5000)